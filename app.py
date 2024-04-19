import gspread
import os
import openpyxl
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
load_dotenv()

wb = openpyxl.Workbook()
sheet1 = wb.active

scope = ['https://www.googleapis.com/auth/spreadsheets']
credentials = ServiceAccountCredentials.from_json_keyfile_name("nutragenanalytics-eae0615deabf.json", scope)
client = gspread.authorize(credentials)

spreadsheet_id = os.getenv("spreadsheet_id")

print(spreadsheet_id)

sheet_name = 'R%'

sheet = client.open_by_key(spreadsheet_id).worksheet(sheet_name)

row_index_data = 44
row_index_74 = 74
row_index = 79
row_index_80 = 80
row_index_81 = 81
row_index_82 = 82
row_index_83 = 83

row = 2
row1 = 32
row2 = 37
row3 = 38
row4 = 39
row5 = 40
row6 = 41

row_ = 2
row11 = 32
row21 = 37
row31 = 38
row41 = 39
row51 = 40
row61 = 41

row_i = 86
row111 = 116
row212 = 121
row311 = 122
row412 = 123
row511 = 124
row611 = 125

row_a = 128
row_b = 158
row_c = 163
row_d = 164
row_e = 165
row_f = 166
row_g = 167

row_a1 = 170
row_b1 = 200
row_c1 = 205
row_d1 = 206
row_e1 = 207
row_f1 = 208
row_g1 = 209

row_a11 = 212
row_b11 = 242
row_c11 = 247
row_d11 = 248
row_e11 = 249
row_f11 = 250
row_g11 = 251

row_a111 = 254
row_b111 = 284
row_c111 = 289
row_d111 = 290
row_e111 = 291
row_f111 = 292
row_g111 = 293

row_a1111 = 506
row_b1111 = 511
row_c1111 = 514
row_d1111 = 515
row_e1111 = 516
row_f1111 = 517
row_g1111 = 518

row_data1 = sheet.row_values(row_index_data)
row_data = sheet.row_values(row_index)
row_data2 = sheet.row_values(row_index_74)
row_data3 = sheet.row_values(row_index_80)
row_data4 = sheet.row_values(row_index_81)
row_data5 = sheet.row_values(row_index_82)
row_data6 = sheet.row_values(row_index_83)


row_data11 = sheet.row_values(row)
row_datapu = sheet.row_values(row1)
row_data23 = sheet.row_values(row2)
row_data34 = sheet.row_values(row3)
row_data45 = sheet.row_values(row4)
row_data56 = sheet.row_values(row5)
row_data67 = sheet.row_values(row6)

row_data111 = sheet.row_values(row_a)
row_datapu1 = sheet.row_values(row_b)
row_data231 = sheet.row_values(row_c)
row_data341 = sheet.row_values(row_d)
row_data451 = sheet.row_values(row_e)
row_data561 = sheet.row_values(row_f)
row_data671 = sheet.row_values(row_g)

row_data1111 = sheet.row_values(row_i)
row_datapu11 = sheet.row_values(row111)
row_data2311 = sheet.row_values(row212)
row_data3411 = sheet.row_values(row311)
row_data4511 = sheet.row_values(row412)
row_data5611 = sheet.row_values(row511)
row_data6711 = sheet.row_values(row611)

row_data11111 = sheet.row_values(row_a1)
row_datapu111 = sheet.row_values(row_b1)
row_data23111 = sheet.row_values(row_c1)
row_data34111 = sheet.row_values(row_d1)
row_data45111 = sheet.row_values(row_e1)
row_data56111 = sheet.row_values(row_f1)
row_data67111 = sheet.row_values(row_g1)

row_data111111 = sheet.row_values(row_a11)
row_datapu1111 = sheet.row_values(row_b11)
row_data231111 = sheet.row_values(row_c11)
row_data341111 = sheet.row_values(row_d11)
row_data451111 = sheet.row_values(row_e11)
row_data561111 = sheet.row_values(row_f11)
row_data671111 = sheet.row_values(row_g11)

row_data1111111 = sheet.row_values(row_a111)
row_datapu11111 = sheet.row_values(row_b111)
row_data2311111 = sheet.row_values(row_c111)
row_data3411111 = sheet.row_values(row_d111)
row_data4511111 = sheet.row_values(row_e111)
row_data5611111 = sheet.row_values(row_f111)
row_data6711111 = sheet.row_values(row_g111)

row_data11111111 = sheet.row_values(row_a1111)
row_datapu111111 = sheet.row_values(row_b1111)
row_data23111111 = sheet.row_values(row_c1111)
row_data34111111 = sheet.row_values(row_d1111)
row_data45111111 = sheet.row_values(row_e1111)
row_data56111111 = sheet.row_values(row_f1111)
row_data67111111 = sheet.row_values(row_g1111)




my_list11 = [item.replace('\xa0', '') for item in row_data11]
my_list21 = [item.replace('\xa0', '') for item in row_datapu]
my_list31 = [item.replace('\xa0', '') for item in row_data23]
my_list41 = [item.replace('\xa0', '') for item in row_data34]
my_list51 = [item.replace('\xa0', '') for item in row_data45]
my_list61 = [item.replace('\xa0', '') for item in row_data56]
my_list71 = [item.replace('\xa0', '') for item in row_data67]

my_list1 = [item.replace('\xa0', '') for item in row_data1]
my_list2 = [item.replace('\xa0', '') for item in row_data]
my_list3 = [item.replace('\xa0', '') for item in row_data2]
my_list4 = [item.replace('\xa0', '') for item in row_data3]
my_list5 = [item.replace('\xa0', '') for item in row_data4]
my_list6 = [item.replace('\xa0', '') for item in row_data5]
my_list7 = [item.replace('\xa0', '') for item in row_data6]

my_list111 = [item.replace('\xa0', '') for item in row_data111]
my_list211 = [item.replace('\xa0', '') for item in row_datapu1]
my_list311 = [item.replace('\xa0', '') for item in row_data231]
my_list411 = [item.replace('\xa0', '') for item in row_data341]
my_list511 = [item.replace('\xa0', '') for item in row_data451]
my_list611 = [item.replace('\xa0', '') for item in row_data561]
my_list711 = [item.replace('\xa0', '') for item in row_data671]

my_list1111 = [item.replace('\xa0', '') for item in row_data1111]
my_list2111 = [item.replace('\xa0', '') for item in row_datapu11]
my_list3111 = [item.replace('\xa0', '') for item in row_data2311]
my_list4111 = [item.replace('\xa0', '') for item in row_data3411]
my_list5111 = [item.replace('\xa0', '') for item in row_data4511]
my_list6111 = [item.replace('\xa0', '') for item in row_data5611]
my_list7111 = [item.replace('\xa0', '') for item in row_data6711]

my_list11111 = [item.replace('\xa0', '') for item in row_data11111]
my_list21111 = [item.replace('\xa0', '') for item in row_datapu111]
my_list31111 = [item.replace('\xa0', '') for item in row_data23111]
my_list41111 = [item.replace('\xa0', '') for item in row_data34111]
my_list51111 = [item.replace('\xa0', '') for item in row_data45111]
my_list61111 = [item.replace('\xa0', '') for item in row_data56111]
my_list71111 = [item.replace('\xa0', '') for item in row_data67111]

my_list111111 = [item.replace('\xa0', '') for item in row_data111111]
my_list211111 = [item.replace('\xa0', '') for item in row_datapu1111]
my_list311111 = [item.replace('\xa0', '') for item in row_data231111]
my_list411111 = [item.replace('\xa0', '') for item in row_data341111]
my_list511111 = [item.replace('\xa0', '') for item in row_data451111]
my_list611111 = [item.replace('\xa0', '') for item in row_data561111]
my_list711111 = [item.replace('\xa0', '') for item in row_data671111]

my_list1111111 = [item.replace('\xa0', '') for item in row_data1111111]
my_list2111111 = [item.replace('\xa0', '') for item in row_datapu11111]
my_list3111111 = [item.replace('\xa0', '') for item in row_data2311111]
my_list4111111 = [item.replace('\xa0', '') for item in row_data3411111]
my_list5111111 = [item.replace('\xa0', '') for item in row_data4511111]
my_list6111111 = [item.replace('\xa0', '') for item in row_data5611111]
my_list7111111 = [item.replace('\xa0', '') for item in row_data6711111]

my_list11111111 = [item.replace('\xa0', '') for item in row_data11111111]
my_list21111111 = [item.replace('\xa0', '') for item in row_datapu111111]
my_list31111111 = [item.replace('\xa0', '') for item in row_data23111111]
my_list41111111 = [item.replace('\xa0', '') for item in row_data34111111]
my_list51111111 = [item.replace('\xa0', '') for item in row_data45111111]
my_list61111111 = [item.replace('\xa0', '') for item in row_data56111111]
my_list71111111 = [item.replace('\xa0', '') for item in row_data67111111]

for i in range(len(my_list1)):
    sheet1.cell(row=2, column=i+1, value=my_list1[i])
for i in range(len(my_list2)):
    sheet1.cell(row=4, column=i+1, value=my_list2[i])
for i in range(len(my_list3)):
    sheet1.cell(row=3, column=i+1, value=my_list3[i])
for i in range(len(my_list4)):
    sheet1.cell(row=5, column=i+1, value=my_list4[i])
for i in range(len(my_list5)):
    sheet1.cell(row=6, column=i+1, value=my_list5[i])
for i in range(len(my_list6)):
    sheet1.cell(row=7, column=i+1, value=my_list6[i])
for i in range(len(my_list7)):
    sheet1.cell(row=8, column=i+1, value=my_list7[i])


for i in range(len(my_list11)):
    sheet1.cell(row=10, column=i+1, value=my_list11[i])
for i in range(len(my_list21)):
    sheet1.cell(row=11, column=i+1, value=my_list21[i])
for i in range(len(my_list31)):
    sheet1.cell(row=12, column=i+1, value=my_list31[i])
for i in range(len(my_list41)):
    sheet1.cell(row=13, column=i+1, value=my_list41[i])
for i in range(len(my_list51)):
    sheet1.cell(row=14, column=i+1, value=my_list51[i])
for i in range(len(my_list61)):
    sheet1.cell(row=15, column=i+1, value=my_list61[i])
for i in range(len(my_list71)):
    sheet1.cell(row=16, column=i+1, value=my_list71[i])

for i in range(len(my_list111)):
    sheet1.cell(row=18, column=i+1, value=my_list111[i])
for i in range(len(my_list211)):
    sheet1.cell(row=19, column=i+1, value=my_list211[i])
for i in range(len(my_list311)):
    sheet1.cell(row=20, column=i+1, value=my_list311[i])
for i in range(len(my_list411)):
    sheet1.cell(row=21, column=i+1, value=my_list411[i])
for i in range(len(my_list511)):
    sheet1.cell(row=22, column=i+1, value=my_list511[i])
for i in range(len(my_list611)):
    sheet1.cell(row=23, column=i+1, value=my_list611[i])
for i in range(len(my_list711)):
    sheet1.cell(row=24, column=i+1, value=my_list711[i])

for i in range(len(my_list1111)):
    sheet1.cell(row=26, column=i+1, value=my_list1111[i])
for i in range(len(my_list2111)):
    sheet1.cell(row=27, column=i+1, value=my_list2111[i])
for i in range(len(my_list3111)):
    sheet1.cell(row=28, column=i+1, value=my_list3111[i])
for i in range(len(my_list4111)):
    sheet1.cell(row=29, column=i+1, value=my_list4111[i])
for i in range(len(my_list5111)):
    sheet1.cell(row=30, column=i+1, value=my_list5111[i])
for i in range(len(my_list6111)):
    sheet1.cell(row=31, column=i+1, value=my_list6111[i])
for i in range(len(my_list7111)):
    sheet1.cell(row=32, column=i+1, value=my_list7111[i])

for i in range(len(my_list11111)):
    sheet1.cell(row=34, column=i+1, value=my_list11111[i])
for i in range(len(my_list21111)):
    sheet1.cell(row=35, column=i+1, value=my_list21111[i])
for i in range(len(my_list31111)):
    sheet1.cell(row=36, column=i+1, value=my_list31111[i])
for i in range(len(my_list41111)):
    sheet1.cell(row=37, column=i+1, value=my_list41111[i])
for i in range(len(my_list51111)):
    sheet1.cell(row=38, column=i+1, value=my_list51111[i])
for i in range(len(my_list61111)):
    sheet1.cell(row=39, column=i+1, value=my_list61111[i])
for i in range(len(my_list71111)):
    sheet1.cell(row=40, column=i+1, value=my_list71111[i])

for i in range(len(my_list111111)):
    sheet1.cell(row=42, column=i+1, value=my_list111111[i])
for i in range(len(my_list211111)):
    sheet1.cell(row=43, column=i+1, value=my_list211111[i])
for i in range(len(my_list311111)):
    sheet1.cell(row=44, column=i+1, value=my_list311111[i])
for i in range(len(my_list411111)):
    sheet1.cell(row=45, column=i+1, value=my_list411111[i])
for i in range(len(my_list511111)):
    sheet1.cell(row=46, column=i+1, value=my_list511111[i])
for i in range(len(my_list611111)):
    sheet1.cell(row=47, column=i+1, value=my_list611111[i])
for i in range(len(my_list711111)):
    sheet1.cell(row=48, column=i+1, value=my_list711111[i])

for i in range(len(my_list1111111)):
    sheet1.cell(row=50, column=i+1, value=my_list1111111[i])
for i in range(len(my_list2111111)):
    sheet1.cell(row=51, column=i+1, value=my_list2111111[i])
for i in range(len(my_list3111111)):
    sheet1.cell(row=52, column=i+1, value=my_list3111111[i])
for i in range(len(my_list4111111)):
    sheet1.cell(row=53, column=i+1, value=my_list4111111[i])
for i in range(len(my_list5111111)):
    sheet1.cell(row=54, column=i+1, value=my_list5111111[i])
for i in range(len(my_list6111111)):
    sheet1.cell(row=55, column=i+1, value=my_list6111111[i])
for i in range(len(my_list7111111)):
    sheet1.cell(row=56, column=i+1, value=my_list7111111[i])

for i in range(len(my_list11111111)):
    sheet1.cell(row=58, column=i+1, value=my_list11111111[i])
for i in range(len(my_list21111111)):
    sheet1.cell(row=59, column=i+1, value=my_list21111111[i])
for i in range(len(my_list31111111)):
    sheet1.cell(row=60, column=i+1, value=my_list31111111[i])
for i in range(len(my_list41111111)):
    sheet1.cell(row=61, column=i+1, value=my_list41111111[i])
for i in range(len(my_list51111111)):
    sheet1.cell(row=62, column=i+1, value=my_list51111111[i])
for i in range(len(my_list61111111)):
    sheet1.cell(row=63, column=i+1, value=my_list61111111[i])
for i in range(len(my_list71111111)):
    sheet1.cell(row=64, column=i+1, value=my_list71111111[i])

def get_column_label(i):
    if i < 26:
        return chr(65 + i)
    else:
        div = i // 26
        mod = i % 26
        if mod == 0:
            return get_column_label(div - 1) + 'Z'
        else:
            return get_column_label(div - 1) + get_column_label(mod)

wb.save("sheet.xlsx")

def CopyFromExcInGsh():
    client = gspread.authorize(credentials)
    viID = os.getenv('viID')

    spreadsheet = client.open_by_key(viID)
    worksheet = spreadsheet.worksheet('cosmo')

    df = pd.read_excel("sheet.xlsx")
    data_list = df.values.tolist()
    num_cols = len(data_list[0])

    cell_list = worksheet.range('A1:' + get_column_label(num_cols-1) + str(len(data_list)))
    for cell in cell_list:
        row = (cell.row - 1) if (cell.row - 1) < len(data_list) else -1
        col = (cell.col - 1) if (cell.col - 1) < num_cols else -1
        if row != -1 and col != -1:
            value = data_list[row][col]
            if pd.notna(value):
                cell.value = str(value)

    worksheet.update_cells(cell_list)
    print("Данные загружены")

CopyFromExcInGsh()
